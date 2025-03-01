import os
import csv
import pandas as pd
from datetime import datetime
import openpyxl
from functools import cmp_to_key
from config import *

class Order:
    def __init__(self, PO, vendor, shipToLocation, ASIN, externalId, externalIdType, modelNumber, title, availability, windowType, windowStart, windowEnd, expectedDate, quantityRequested, expectedQuantity, unitCost, currencyCode):
        qty = 1 if modelNumber.split('-')[-1] == 'EACH' else int((modelNumber).split('-')[-1])
        self.PO = PO
        self.vendor = vendor
        self.shipToLocation = shipToLocation
        self.ASIN = ASIN
        self.externalId = externalId
        self.externalIdType = externalIdType
        self.modelNumber = modelNumber
        self.itemNumber = None
        self.title = title
        self.availability = availability
        self.windowType = windowType
        self.windowStart = windowStart
        self.windowEnd = windowEnd
        self.expectedDate = expectedDate
        self.quantityRequested = int(quantityRequested)
        self.qtyInEach = qty * int(quantityRequested) if '-' in modelNumber else 1
        self.expectedQuantity = int(expectedQuantity)
        self.unitCost = float(unitCost)
        self.currencyCode = currencyCode
        self.uomCode = None
        self.totalPrice = self.unitCost * self.quantityRequested

    def __str__(self):
        return 'PO: {}, Item Number: {}, UOM Code: {}, Qty: {}, Total: {}'.format(self.PO, self.modelNumber, self.uomCode, self.quantityRequested, self.totalPrice)

def getTimestamp():
    now = datetime.now()
    return datetime.strftime(now, "%m%d%Y%H%M%S")

def getCurrentime():
    return datetime.now()

def getFileModifiedDate(filepath):
    return datetime.fromtimestamp(os.path.getmtime(filepath))

def getDaysDifferent(currentTime, timestamp):
    return (currentTime - timestamp).days

def getUOMMasterData(inputFilepath):
    mapped = {}

    try:
        with open (inputFilepath, mode='r') as file:
            content = csv.reader(file)
            for line in content:
                if (len(line) == 3):
                    mapped['{}-{}'.format(line[1], line[2])] = {
                        "modelNumber": line[0],
                        "sku": line[1],
                        "uomQty": line[2]
                    }
    except:
        print('*** Error: Failed to read input file for UOM Master. Please make sure filename is valid. ***')
        return {}

    return mapped

def getUOMMasterDataWithExcelFormat(inputFilepath):
    mapped = {}
    message = None
    
    try:
        age = getDaysDifferent(getCurrentime(), getFileModifiedDate(inputFilepath))
        message = 'UOM master file was updated {} days ago.'.format(age)

        workbook = openpyxl.load_workbook(inputFilepath)
        sheet = workbook.active
        for r in range(2, sheet.max_row+1):
            modelNumber = None
            sku = None
            uomQty = None

            for c in range(1, sheet.max_column+1):
                data = sheet.cell(row=r, column=c).value
                if c == 1:
                    modelNumber = str(data)
                if c == 2:
                    sku = str(data)
                if c == 3:
                    uomQty = int(data)

                if (modelNumber and sku and uomQty):
                    mapped['{}-{}'.format(sku, uomQty)] = {
                        "modelNumber": modelNumber,
                        "sku": sku,
                        "uomQty": uomQty
                    }
    except:
        print('*** Error: Failed to read input file for UOM Master. Please make sure filename is valid. ***')
        return {}, message
    # print(mapped)
    return mapped, message

def getInventoryAndPriceMasterData(inputFilepath):
    age = getDaysDifferent(getCurrentime(), getFileModifiedDate(inputFilepath))
    message = 'Inventory master file was updated {} days ago.'.format(age)

    mapped = {}

    try:
        workbook = openpyxl.load_workbook(inputFilepath) # #, Item No., Item Desc., Available Qty
        sheet = workbook.active
        for r in range(2, sheet.max_row+1):
            itemNumber = None
            for c in range(1, sheet.max_column+1):
                data = sheet.cell(row=r, column=c).value
                if c == 2: # item number
                    itemNumber = str(data)
                    mapped[itemNumber] = {}
                if c == 4: # stock
                    mapped[itemNumber]["qty"] = data
                if c == 8: # P1000
                    mapped[itemNumber]["p1000"] = data
    except:
        print('*** Error: Failed to read input file for Inventory Master. Please make sure filename is valid. ***')
        return {}, message

    return mapped, message

def getOrdersFromInputfile(filepath):
    orders = []
    
    try:
        with open (filepath, mode='r') as file:
            count = 1
            content = csv.reader(file)
            for line in content:
                if count == 1:
                    count += 1
                    continue
        
                if (len(line) == 17):
                    if not line[6]: # if Model Number not exists
                        continue
                    order = Order(line[0], line[1], line[2], line[3], line[4], line[5], line[6], line[7], line[8], line[9], line[10], line[11], line[12], line[13], line[14], line[15], line[16])
                    orders.append(order)
                count += 1
    except Exception as err:
        message = 'Please check your input batch file: {}'.format(filepath)
        print('*** Error: Failed to read batch input file. Please make sure filename is valid. err: {} ***'.format(err))
        return []

    return orders

# -1 : Rejected, 0 : Rejected, Suggested, 1 : Accepted
def validateOrder(order, sapUnitPrice, sapStock):
    # check if order price match the agreed upon price
    if round(sapUnitPrice, 2) > round(order.unitCost, 2):
        return -1
    # check if item is not out of stock
    elif sapStock < (order.qtyInEach):
        return -1
    # check if total price is over or equal to $30
    # elif order.totalPrice < 30:
    #     return 0
    else:
        return 1

def processAmazonVendorCentralOrders(orders, uomMaster, qtyPriceMaster):
    acceptedOrders = []
    rejectedOrders = []
    suggestedOrders = []

    for order in orders:
        if order.modelNumber in uomMaster:
            order.uomCode = (uomMaster[order.modelNumber]["modelNumber"]).split('-')[-1]
        else:
            order.uomCode = 'EA'

        if '-' not in order.modelNumber or '-EACH' in order.modelNumber:
            rejectedOrders.append([order.PO, order.modelNumber, order.quantityRequested, order.unitCost, order.uomCode, order.totalPrice, 0, 'Invalid SKU'])
            continue

        order.itemNumber = order.modelNumber if '-' not in order.modelNumber else (uomMaster[order.modelNumber]["sku"] if order.modelNumber in uomMaster else None)
 
        if not order.itemNumber in qtyPriceMaster:
            rejectedOrders.append([order.PO, order.modelNumber, order.quantityRequested, order.unitCost, order.uomCode, order.totalPrice, 0, 'Not in SAP'])
            continue

        sapPpP = qtyPriceMaster[order.itemNumber]['p1000'] if order.itemNumber in qtyPriceMaster else 0
        sapUnitPrice = (sapPpP * order.qtyInEach) / order.quantityRequested
        sapStock = qtyPriceMaster[order.itemNumber]['qty'] if order.itemNumber in qtyPriceMaster else 0

        validation = validateOrder(order, sapUnitPrice, sapStock)
        if validation == -1:
            rejectedOrders.append([order.PO, order.modelNumber, order.quantityRequested, order.unitCost, order.uomCode, order.totalPrice, sapUnitPrice, 'Price'])
        # elif validation == 0:
        #     rejectedOrders.append([order.PO, order.modelNumber, order.quantityRequested, order.unitCost, order.uomCode, order.totalPrice, sapUnitPrice, '< $30'])
        #     suggestedOrders.append([order.PO, order.modelNumber, order.quantityRequested, order.unitCost, order.uomCode, order.totalPrice, sapUnitPrice, '< $30'])
        elif validation == 1:
            pricePerPiece = order.totalPrice / order.qtyInEach
            acceptedOrders.append([order.itemNumber, '', order.uomCode, order.quantityRequested, pricePerPiece, '', '', order.PO, order.modelNumber, order.quantityRequested, order.unitCost, order.uomCode, order.totalPrice, sapUnitPrice])
        else:
            pass
    
    return acceptedOrders, rejectedOrders, suggestedOrders

def validateInputFilename(filename):
    cleaned = filename
    if '/' in filename:
        cleaned = filename.split('/')[-1]

    if '.csv' not in cleaned:
        cleaned = cleaned + '.csv'

    return USER_DOWNLOADS + cleaned

def getUOMMasterFilepath():
    return os.path.join(ASSETS_BASE_DIR, UOM_MASTER_FILENAME)

def getQtyPriceMasterFilepath():
    return os.path.join(ASSETS_BASE_DIR, QTY_PRICE_MASTER_FILENAME)

def writeLog(timestamp, status):
    path = os.path.join(ASSETS_BASE_DIR, LOGS_FILENAME)
    user = os.getenv('COMPUTERNAME')
    try:
        with open(path, 'a') as file:
            file.write('USR;{} | IN;{} | SUCCESS;{} | ERR;{} | WARNING;{} | WARN;{} | OUT;{} | VER;{} | TS;{}\n'.format(user, status["inputFilename"], status["success"], status["errorMessage"], status["warning"], status["warningMessage"], status["outputFilename"], APP_VERSION, timestamp))
    except:
        print('*** Error: Failed to write to logs. ***')

def sortOrders(a, b):
    if a[4] == 'CASE' and (b[4] == 'BOX' or b[4] == 'EA' or b[4] == None):
        return -1
    elif a[4] == 'BOX' and (b[4] == 'EA' or b[4] == None):
        return -1
    else:
        return 1
    
def sortOrdersForAccepted(a, b):
    if a[2] == 'CASE' and (b[2] == 'BOX' or b[2] == 'EA' or b[2] == None):
        return -1
    elif a[2] == 'BOX' and (b[2] == 'EA' or b[2] == None):
        return -1
    else:
        return 1
    
# def sortOrdersForRejectedAndSuggested(a, b):
#     if a[4] == 'CASE' and (b[4] == 'BOX' or b[4] == 'EA'):
#         return -1
#     elif a[4] == 'BOX' and b[4] == 'EA':
#         return -1
#     else:
#         return 1

def processResult(inputFilepath):
    input = validateInputFilename(inputFilepath)
    timestamp = getTimestamp()

    response = {
        "success": True,
        "errorMessage": "",
        "inputFilename": inputFilepath,
        "outputFilename": "",
        "warning": None,
        "warningMessage": None,
    }

    uomMasterFilepath = getUOMMasterFilepath()
    qtyPriceMasterFilepath = getQtyPriceMasterFilepath()

    uomMaster, uomMasterMsg = getUOMMasterDataWithExcelFormat(uomMasterFilepath)
    qtyPriceMaster, qtyPriceMsg = getInventoryAndPriceMasterData(qtyPriceMasterFilepath)
    orders = getOrdersFromInputfile(input)

    acceptedOrders, rejectedOrders, suggestedOrders = processAmazonVendorCentralOrders(orders, uomMaster, qtyPriceMaster)
   
    if not acceptedOrders and not rejectedOrders:
        response["success"] = False
        response["errorMessage"] = "Please try again or contact admin."
        writeLog(timestamp, response)
        return response
    
    totalPOPrice = 0
    for order in acceptedOrders:
        totalPOPrice += order[12]

    acceptedOrders.sort(key=cmp_to_key(sortOrdersForAccepted))
    rejectedOrders.sort(key=cmp_to_key(sortOrders))
    suggestedOrders.sort(key=cmp_to_key(sortOrders))

    outputFilename = 'po_output_{}.xlsx'.format(timestamp)
    outputFilepath = OUTPUT_DIR + outputFilename

    acceptedDF = pd.DataFrame(acceptedOrders, columns=['SKU', 'Desc', 'UOM', 'QTY', 'PpP', '', '', 'PO', 'Item Number', 'Qty', 'Amazon Price', 'UOM', 'Total Price', 'Bazic Price']) # columns=['PO', 'Item Number', 'Qty', 'Unit Cost', 'UOM', 'Total Price', 'SAP Unit Cost']
    acceptedDF.index = acceptedDF.index + 1

    rejectedDF = pd.DataFrame(rejectedOrders, columns=['PO', 'Item Number', 'Qty', 'Amazon Price', 'UOM', 'Total Price', 'Bazic Price', 'Reason'])
    rejectedDF.index = rejectedDF.index + 1

    suggestedDF = pd.DataFrame(suggestedOrders, columns=['PO', 'Item Number', 'Qty', 'Amazon Price', 'UOM', 'Total Price', 'Bazic Price', 'Reason'])
    suggestedDF.index = suggestedDF.index + 1

    with pd.ExcelWriter(outputFilepath, engine='xlsxwriter') as writer:
        acceptedDF.to_excel(writer, sheet_name='Accepted', startrow=2, startcol=0)
        rejectedDF.to_excel(writer, sheet_name='Rejected', startrow=2, startcol=0)
        suggestedDF.to_excel(writer, sheet_name='Optional', startrow=2, startcol=0)

        worksheet = writer.sheets['Accepted']
        worksheet.write(0, 0, "Total PO Price: ${:.2f}".format(totalPOPrice))

    response["outputFilename"] = outputFilepath
    writeLog(timestamp, response)

    return response